from flask import Flask, render_template, send_file, request, redirect, url_for
from src.model import solve_schedule, DAYS, LOCATIONS, HOURS
from src.db.data_access import (
    load_employees, add_employee, delete_employee, update_employee,
    get_employee_by_id, set_availability, get_availability_for_employee,
    set_location_preferences, get_preferences_for_employee,
    set_mopper_status, load_saved_schedule
)
import openpyxl
from openpyxl.styles import Font, PatternFill
import io

def format_hour(h):
    """Converts a 24-hour int (e.g. 11, 17) to a 12-hour label (e.g. '11 A.M.', '5 P.M.')"""
    period = "A.M." if h < 12 else "P.M."
    display_hour = h % 12
    if display_hour == 0:
        display_hour = 12
    return f"{display_hour} {period}"

app = Flask(__name__)

app.jinja_env.filters["format_hour"] = format_hour

def build_shift_blocks(schedule):
    """Collapses hourly assignments into contiguous shift blocks.
    Returns blocks[location][day] = list of {employee, start, end} dicts."""
    # Derive the set of employees actually present in this schedule,
    # since EMPLOYEES is no longer a fixed module-level list.
    employees_in_schedule = {
        e
        for d in schedule
        for loc in schedule[d]
        for h in schedule[d][loc]
        for e in schedule[d][loc][h]
    }

    blocks = {loc: {d: [] for d in DAYS} for loc in LOCATIONS}
    for loc in LOCATIONS:
        for d in DAYS:
            for e in employees_in_schedule:
                hours_here = [h for h in HOURS if e in schedule[d][loc][h]]
                if not hours_here:
                    continue
                start = hours_here[0]
                prev = hours_here[0]
                for h in hours_here[1:] + [None]:
                    if h is None or h != prev + 1:
                        blocks[loc][d].append({"employee": e, "start": start, "end": prev + 1})
                        if h is not None:
                            start = h
                    if h is not None:
                        prev = h
    return blocks

def compute_gantt_layout(blocks):
    """Adds lane/position info to each block for CSS-based rendering.
    Time flows left-to-right (x-axis), days are stacked top-to-bottom (rows)."""
    day_start, day_end = HOURS[0], HOURS[-1] + 1
    total_span = day_end - day_start

    layout = {loc: {d: [] for d in blocks[list(blocks.keys())[0]]} for loc in blocks}
    for loc in blocks:
        for d in blocks[loc]:
            day_blocks = sorted(blocks[loc][d], key=lambda b: b["start"])

            # Group into clusters of overlapping blocks (still needed for lane assignment)
            clusters = []
            current_cluster = []
            current_end = None
            for b in day_blocks:
                if current_cluster and b["start"] < current_end:
                    current_cluster.append(b)
                    current_end = max(current_end, b["end"])
                else:
                    if current_cluster:
                        clusters.append(current_cluster)
                    current_cluster = [b]
                    current_end = b["end"]
            if current_cluster:
                clusters.append(current_cluster)

            for cluster in clusters:
                lanes_end = []
                for b in cluster:
                    placed = False
                    for lane_idx, lane_end in enumerate(lanes_end):
                        if b["start"] >= lane_end:
                            lanes_end[lane_idx] = b["end"]
                            b["lane"] = lane_idx
                            placed = True
                            break
                    if not placed:
                        b["lane"] = len(lanes_end)
                        lanes_end.append(b["end"])

                num_lanes = len(lanes_end)
                for b in cluster:
                    b["num_lanes"] = num_lanes
                    # SWAPPED: time now drives horizontal position/width
                    b["left_pct"] = (b["start"] - day_start) / total_span * 100
                    b["width_pct"] = (b["end"] - b["start"]) / total_span * 100
                    # SWAPPED: lane now drives vertical position/height
                    b["top_pct"] = b["lane"] / num_lanes * 100
                    b["height_pct"] = 100 / num_lanes

            layout[loc][d] = day_blocks
    return layout

@app.route("/")
def index():
    saved = load_saved_schedule()
    if saved is None:
        return render_template("no_schedule.html")
    blocks = build_shift_blocks(saved["schedule"])
    gantt = compute_gantt_layout(blocks)
    return render_template(
        "schedule.html",
        gantt=gantt,
        days=DAYS,
        locations=LOCATIONS,
        hours=HOURS,
        generated_at=saved["generated_at"]
    )

@app.route("/generate", methods=["GET", "POST"])
def generate():
    if request.method == "GET":
        return render_template("generate.html")
    
     # POST: actually run the solver
    result = solve_schedule()
    if result["feasible"]:
        return redirect(url_for("index"))
    else:
        return render_template("infeasible.html", diagnostics=result["diagnostics"])

@app.route("/export")
def export():
    saved = load_saved_schedule()
    if saved is None:
        return "No schedule has been generated yet.", 400

    blocks = build_shift_blocks(saved["schedule"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)

    for loc in LOCATIONS:
        ws = wb.create_sheet(title=loc[:31])

        # Only include employees who actually worked this location this week
        employees_here = sorted({
            b["employee"]
            for d in DAYS
            for b in blocks[loc][d]
        })

        ws.append(["Employee"] + DAYS)
        for cell in ws[1]:
            cell.font = header_font

        for e in employees_here:
            row = [e]
            for d in DAYS:
                day_blocks = [b for b in blocks[loc][d] if b["employee"] == e]
                if day_blocks:
                    row.append(", ".join(f"{format_hour(b['start'])}–{format_hour(b['end'])}" for b in day_blocks))
                else:
                    row.append("")
            ws.append(row)

        # Grey out empty cells
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if not cell.value:
                    cell.fill = grey_fill

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="schedule.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/manage")
def manage():
    employees = load_employees()  # list of (id, name, max_hours)
    return render_template("manage.html", employees=employees)

@app.route("/manage/add", methods=["POST"])
def manage_add():
    name = request.form["name"].strip()
    max_hours = int(request.form.get("max_hours", 40))
    if name:
        add_employee(name, max_hours)
    return redirect(url_for("manage"))

@app.route("/manage/delete/<int:employee_id>", methods=["POST"])
def manage_delete(employee_id):
    delete_employee(employee_id)
    return redirect(url_for("manage"))

@app.route("/manage/edit/<int:employee_id>", methods=["GET", "POST"])
def manage_edit(employee_id):
    if request.method == "POST":
        name = request.form["name"].strip()
        max_hours = int(request.form["max_hours"])
        update_employee(employee_id, name, max_hours)

        # Rebuild availability from submitted checkboxes
        selected = request.form.getlist("availability")  # list of "Mon-11" style strings
        availability_set = set()
        for entry in selected:
            day, hour = entry.split("-")
            availability_set.add((day, int(hour)))
        set_availability(employee_id, availability_set)

        # Rebuild preferences
        preferences = {}
        for loc in LOCATIONS:
            score = request.form.get(f"pref_{loc}", 0)
            preferences[loc] = int(score)
        set_location_preferences(employee_id, preferences)

        is_mopper = 1 if request.form.get("is_mopper") else 0
        set_mopper_status(employee_id, is_mopper)

        return redirect(url_for("manage"))

    employee = get_employee_by_id(employee_id)
    current_availability = get_availability_for_employee(employee_id)
    current_preferences = get_preferences_for_employee(employee_id)

    return render_template(
        "edit_employee.html",
        employee=employee,
        days=DAYS,
        hours=HOURS,
        locations=LOCATIONS,
        current_availability=current_availability,
        current_preferences=current_preferences
    )

@app.route("/staffing", methods=["GET", "POST"])
def staffing():
    from src.db.data_access import load_staffing_requirements_grid, set_staffing_requirement
    if request.method == "POST":
        for loc in LOCATIONS:
            for d in DAYS:
                for h in HOURS:
                    min_val = int(request.form.get(f"min_{loc}_{d}_{h}", 0))
                    set_staffing_requirement(loc, d, h, min_val)
        return redirect(url_for("staffing"))

    requirements = load_staffing_requirements_grid()
    return render_template(
        "staffing.html",
        requirements=requirements,
        days=DAYS,
        locations=LOCATIONS,
        hours=HOURS
    )

if __name__ == "__main__":
    app.run(debug=True, port=5001)